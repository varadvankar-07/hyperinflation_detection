import pandas as pd
import numpy as np
from datetime import datetime


def create_comprehensive_inflation_dataset():
    """
    Create a comprehensive inflation dataset suitable for data science projects
    with real historical data for major economies
    """

    # Years from 1980 to 2024
    years = list(range(1980, 2025))

    # Comprehensive inflation data for 15 major economies
    inflation_data = [
        {
            'Country': 'United States',
            'Region': 'North America',
            'Income_Level': 'High Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                13.5, 10.3, 6.2, 3.2, 4.3, 3.6, 1.9, 3.6, 4.1, 4.8,
                5.4, 4.2, 3.0, 3.0, 2.6, 2.8, 2.9, 2.3, 1.6, 2.2,
                3.4, 2.8, 1.6, 2.3, 2.7, 3.4, 3.2, 2.9, 3.8, -0.4,
                1.6, 3.2, 2.1, 1.5, 1.6, 0.1, 1.3, 2.1, 2.4, 1.8,
                1.2, 4.7, 8.0, 4.1, 3.4
            ])}
        },
        {
            'Country': 'United Kingdom',
            'Region': 'Europe',
            'Income_Level': 'High Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                18.0, 11.9, 8.6, 4.6, 5.0, 6.1, 3.4, 4.2, 4.9, 7.8,
                9.5, 5.9, 3.7, 1.6, 2.5, 3.4, 2.4, 3.1, 3.4, 1.5,
                2.9, 1.8, 1.7, 2.9, 3.0, 2.8, 3.2, 4.3, 4.0, 2.2,
                3.3, 4.5, 2.8, 2.6, 1.5, 0.0, 0.7, 2.7, 2.5, 1.8,
                0.9, 2.6, 9.1, 7.3, 2.5
            ])}
        },
        {
            'Country': 'Germany',
            'Region': 'Europe',
            'Income_Level': 'High Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                5.4, 6.3, 5.3, 3.3, 2.4, 2.2, -0.1, 0.2, 1.3, 2.8,
                2.7, 3.5, 5.1, 4.5, 2.7, 1.8, 1.4, 1.9, 1.0, 0.6,
                1.4, 1.9, 1.4, 1.0, 1.8, 1.9, 1.8, 2.3, 2.8, 0.2,
                1.2, 2.5, 2.1, 1.6, 0.8, 0.1, 0.4, 1.7, 1.9, 1.4,
                0.4, 3.2, 8.7, 5.9, 2.8
            ])}
        },
        {
            'Country': 'Japan',
            'Region': 'Asia',
            'Income_Level': 'High Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                7.8, 4.9, 2.7, 1.9, 2.3, 2.0, 0.6, 0.1, 0.7, 2.3,
                3.1, 3.3, 1.7, 1.3, 0.7, -0.1, 0.1, 1.7, 0.7, -0.3,
                -0.7, -0.8, -0.9, -0.3, 0.0, -0.3, 0.2, 0.1, 1.4, -1.4,
                -0.7, -0.3, 0.0, 0.3, 2.8, 0.8, -0.1, 0.5, 1.0, 0.5,
                0.0, -0.2, 2.5, 3.3, 2.7
            ])}
        },
        {
            'Country': 'China',
            'Region': 'Asia',
            'Income_Level': 'Upper Middle Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                6.0, 2.4, 1.9, 1.5, 2.8, 9.3, 6.5, 7.3, 18.8, 18.0,
                3.1, 3.4, 6.4, 14.7, 24.1, 17.1, 8.3, 2.8, -0.8, -1.4,
                0.4, 0.7, -0.8, 1.2, 3.9, 1.8, 1.5, 4.8, 5.9, -0.7,
                3.3, 5.4, 2.6, 2.6, 2.0, 1.4, 2.0, 1.6, 2.1, 2.9,
                2.5, 0.9, 2.0, 0.2, 0.4
            ])}
        },
        {
            'Country': 'India',
            'Region': 'Asia',
            'Income_Level': 'Lower Middle Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                11.4, 13.1, 7.9, 11.9, 8.3, 5.6, 8.7, 8.8, 9.4, 6.2,
                9.0, 13.9, 11.8, 6.4, 10.2, 10.2, 9.0, 7.2, 13.2, 4.7,
                4.0, 3.8, 4.3, 3.8, 3.8, 4.2, 5.8, 6.4, 8.3, 10.9,
                12.0, 8.9, 9.3, 10.9, 6.7, 4.9, 4.5, 3.6, 3.4, 4.8,
                6.2, 5.1, 6.7, 5.4, 4.9
            ])}
        },
        {
            'Country': 'Brazil',
            'Region': 'South America',
            'Income_Level': 'Upper Middle Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                110.2, 95.2, 99.7, 211.0, 223.8, 235.1, 65.0, 415.8, 1037.5, 1782.9,
                1476.7, 480.2, 1119.1, 2708.6, 916.5, 22.0, 9.1, 4.3, 1.7, 8.9,
                6.0, 7.7, 12.5, 9.3, 7.6, 5.7, 3.1, 4.5, 5.9, 4.3,
                5.9, 6.5, 5.8, 5.9, 6.4, 10.7, 6.3, 2.9, 3.7, 4.3,
                4.5, 10.1, 5.8, 4.6, 4.8
            ])}
        },
        {
            'Country': 'Canada',
            'Region': 'North America',
            'Income_Level': 'High Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                10.2, 12.5, 10.8, 5.8, 4.3, 4.0, 4.1, 4.4, 4.0, 5.0,
                4.8, 5.6, 1.5, 1.8, 0.2, 2.2, 1.6, 1.6, 1.0, 1.7,
                2.7, 2.5, 2.3, 2.8, 1.9, 2.2, 2.0, 2.1, 2.4, 0.3,
                1.8, 2.9, 1.5, 0.9, 2.0, 1.1, 1.4, 1.6, 2.3, 1.9,
                0.7, 3.4, 6.8, 3.9, 2.7
            ])}
        },
        {
            'Country': 'France',
            'Region': 'Europe',
            'Income_Level': 'High Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                13.6, 13.4, 11.8, 9.6, 7.4, 5.8, 2.7, 3.1, 2.7, 3.6,
                3.4, 3.2, 2.4, 2.1, 1.6, 1.8, 2.0, 1.2, 0.7, 0.5,
                1.7, 1.6, 1.9, 2.1, 2.1, 1.7, 1.7, 1.5, 2.8, 0.1,
                1.5, 2.1, 2.0, 0.9, 0.5, 0.0, 0.2, 1.0, 1.8, 1.1,
                0.5, 1.6, 5.2, 4.9, 2.3
            ])}
        },
        {
            'Country': 'Australia',
            'Region': 'Oceania',
            'Income_Level': 'High Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                10.2, 9.7, 11.1, 10.1, 4.0, 6.7, 9.1, 8.5, 7.3, 7.5,
                7.3, 3.2, 1.0, 1.8, 1.9, 4.6, 2.6, 0.3, 0.9, 1.5,
                4.5, 4.4, 3.0, 2.8, 2.3, 2.7, 3.5, 2.3, 4.4, 1.8,
                2.9, 3.4, 1.8, 2.5, 2.5, 1.5, 1.3, 2.0, 1.9, 1.6,
                0.9, 2.9, 6.6, 5.6, 3.8
            ])}
        },
        {
            'Country': 'Russia',
            'Region': 'Europe',
            'Income_Level': 'Upper Middle Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan,
                np.nan, 144.0, 1353.0, 896.0, 302.0, 197.0, 47.8, 14.8, 27.7, 85.7,
                20.8, 21.5, 15.8, 13.7, 10.9, 12.7, 9.7, 9.0, 14.1, 11.7,
                6.9, 8.4, 5.1, 6.8, 7.8, 15.5, 7.0, 3.7, 2.9, 4.5,
                3.4, 6.7, 13.8, 5.9, 7.4
            ])}
        },
        {
            'Country': 'Mexico',
            'Region': 'North America',
            'Income_Level': 'Upper Middle Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                29.8, 28.7, 98.8, 80.8, 59.2, 63.7, 105.7, 159.2, 51.7, 19.7,
                29.9, 18.8, 11.9, 8.0, 7.1, 52.0, 27.7, 15.7, 18.6, 12.3,
                9.5, 4.4, 5.7, 4.0, 5.4, 3.3, 4.1, 3.8, 6.5, 3.6,
                4.4, 3.8, 3.6, 4.0, 4.1, 2.1, 3.4, 6.8, 4.8, 3.6,
                3.2, 7.4, 7.9, 5.5, 4.6
            ])}
        },
        {
            'Country': 'South Korea',
            'Region': 'Asia',
            'Income_Level': 'High Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                28.7, 21.3, 7.2, 3.4, 2.3, 2.5, 2.8, 3.0, 7.1, 5.7,
                8.6, 9.3, 6.2, 4.8, 6.3, 4.5, 4.9, 4.4, 7.5, 0.8,
                2.3, 4.1, 2.8, 3.5, 3.6, 2.8, 2.2, 2.5, 4.7, 2.8,
                3.0, 4.0, 2.2, 1.3, 1.3, 0.7, 1.0, 1.9, 1.5, 0.4,
                0.5, 2.5, 5.1, 3.6, 2.3
            ])}
        },
        {
            'Country': 'Italy',
            'Region': 'Europe',
            'Income_Level': 'High Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                21.2, 17.8, 16.5, 14.7, 10.8, 9.2, 5.8, 4.7, 5.1, 6.3,
                6.5, 6.3, 5.2, 4.5, 4.0, 5.2, 3.9, 1.8, 1.7, 1.7,
                2.5, 2.8, 2.5, 2.7, 2.2, 2.0, 2.1, 1.8, 3.4, 0.8,
                1.5, 2.8, 3.0, 1.2, 0.2, 0.0, -0.1, 1.2, 1.1, 0.6,
                -0.1, 1.9, 8.2, 5.9, 1.2
            ])}
        },
        {
            'Country': 'Spain',
            'Region': 'Europe',
            'Income_Level': 'High Income',
            'Indicator': 'Consumer Price Index (Annual %)',
            **{year: rate for year, rate in zip(years, [
                15.6, 14.6, 14.4, 12.2, 11.3, 8.8, 8.8, 5.3, 4.8, 6.8,
                6.7, 5.9, 5.9, 4.6, 4.7, 4.7, 3.6, 2.0, 1.8, 2.3,
                3.4, 3.6, 3.1, 3.0, 3.0, 3.4, 3.5, 2.8, 4.1, -0.3,
                1.8, 3.2, 2.4, 1.4, -0.1, -0.5, -0.2, 2.0, 1.7, 0.7,
                -0.3, 3.1, 8.4, 3.5, 3.0
            ])}
        }
    ]

    # Create DataFrame
    df = pd.DataFrame(inflation_data)

    # Reorder columns
    cols = ['Country', 'Region', 'Income_Level', 'Indicator'] + years
    df = df[cols]

    return df


def create_summary_statistics(df):
    """
    Create summary statistics for the dataset
    """
    year_cols = [col for col in df.columns if isinstance(col, int)]

    summary_data = []

    for _, row in df.iterrows():
        country = row['Country']
        region = row['Region']
        income = row['Income_Level']

        # Calculate statistics
        values = row[year_cols].dropna()

        summary_data.append({
            'Country': country,
            'Region': region,
            'Income_Level': income,
            'Mean_Inflation': values.mean(),
            'Median_Inflation': values.median(),
            'Std_Deviation': values.std(),
            'Min_Inflation': values.min(),
            'Max_Inflation': values.max(),
            'Latest_2024': row[2024] if 2024 in row.index else np.nan,
            'Avg_1980s': row[list(range(1980, 1990))].mean(),
            'Avg_1990s': row[list(range(1990, 2000))].mean(),
            'Avg_2000s': row[list(range(2000, 2010))].mean(),
            'Avg_2010s': row[list(range(2010, 2020))].mean(),
            'Avg_2020s': row[list(range(2020, 2025))].mean()
        })

    return pd.DataFrame(summary_data)


def create_regional_analysis(df):
    """
    Create regional analysis
    """
    year_cols = [col for col in df.columns if isinstance(col, int)]

    # Group by region
    regional_data = []

    for region in df['Region'].unique():
        region_df = df[df['Region'] == region]

        for year in year_cols:
            avg_inflation = region_df[year].mean()
            regional_data.append({
                'Region': region,
                'Year': year,
                'Avg_Inflation': avg_inflation
            })

    return pd.DataFrame(regional_data)


def create_decade_analysis(df):
    """
    Create decade-wise analysis
    """
    decades = {
        '1980-1989': list(range(1980, 1990)),
        '1990-1999': list(range(1990, 2000)),
        '2000-2009': list(range(2000, 2010)),
        '2010-2019': list(range(2010, 2020)),
        '2020-2024': list(range(2020, 2025))
    }

    decade_data = []

    for _, row in df.iterrows():
        country = row['Country']
        region = row['Region']

        for decade_name, decade_years in decades.items():
            values = row[decade_years].dropna()

            decade_data.append({
                'Country': country,
                'Region': region,
                'Decade': decade_name,
                'Avg_Inflation': values.mean(),
                'Min_Inflation': values.min(),
                'Max_Inflation': values.max(),
                'Volatility_Std': values.std()
            })

    return pd.DataFrame(decade_data)


def export_to_excel(df, filename='Inflation_Dataset_Data_Science_Project.xlsx'):
    """
    Export comprehensive dataset to Excel with multiple sheets
    """
    print(f"\n{'=' * 80}")
    print("CREATING COMPREHENSIVE INFLATION DATASET FOR DATA SCIENCE PROJECT")
    print(f"{'=' * 80}\n")

    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Sheet 1: Main Dataset
            print("[1/6] Creating main dataset...")
            df.to_excel(writer, sheet_name='Main_Dataset', index=False)

            # Sheet 2: Summary Statistics
            print("[2/6] Calculating summary statistics...")
            summary_df = create_summary_statistics(df)
            summary_df.to_excel(writer, sheet_name='Summary_Statistics', index=False)

            # Sheet 3: Regional Analysis
            print("[3/6] Creating regional analysis...")
            regional_df = create_regional_analysis(df)
            regional_df.to_excel(writer, sheet_name='Regional_Analysis', index=False)

            # Sheet 4: Decade Analysis
            print("[4/6] Creating decade analysis...")
            decade_df = create_decade_analysis(df)
            decade_df.to_excel(writer, sheet_name='Decade_Analysis', index=False)

            # Sheet 5: Recent Trends (2020-2024)
            print("[5/6] Analyzing recent trends...")
            recent_cols = ['Country', 'Region', 2020, 2021, 2022, 2023, 2024]
            recent_df = df[recent_cols].copy()
            recent_df.to_excel(writer, sheet_name='Recent_Trends_2020_2024', index=False)

            # Sheet 6: Metadata and Information
            print("[6/6] Adding metadata...")
            metadata = pd.DataFrame({
                'Field': [
                    'Dataset Name',
                    'Description',
                    'Time Period',
                    'Number of Countries',
                    'Number of Years',
                    'Total Data Points',
                    'Indicator',
                    'Source Type',
                    'Regions Covered',
                    'Income Levels',
                    'Missing Data',
                    'Created Date',
                    'Use Case',
                    'Column Description - Country',
                    'Column Description - Region',
                    'Column Description - Income_Level',
                    'Column Description - Indicator',
                    'Column Description - Years'
                ],
                'Value': [
                    'Global Inflation Dataset (1980-2024)',
                    'Comprehensive historical inflation rates for major world economies',
                    '1980-2024 (45 years)',
                    str(len(df)),
                    '45',
                    str(len(df) * 45),
                    'Consumer Price Index (Annual %)',
                    'Historical economic data',
                    'North America, South America, Europe, Asia, Oceania',
                    'High Income, Upper Middle Income, Lower Middle Income',
                    'Minimal (mainly early years for Russia)',
                    datetime.now().strftime('%Y-%m-%d'),
                    'Data Science, Machine Learning, Time Series Analysis, Economic Research',
                    'Name of the country',
                    'Geographic region of the country',
                    'World Bank income classification',
                    'Type of inflation measure',
                    'Annual inflation rate percentage for each year'
                ]
            })
            metadata.to_excel(writer, sheet_name='Metadata', index=False)

            # Format all sheets
            workbook = writer.book
            from openpyxl.styles import Font, PatternFill, Alignment

            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Format header row
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass

                    adjusted_width = min(max_length + 3, 25)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        print(f"\n{'=' * 80}")
        print(f"✓ SUCCESS! Excel file created: {filename}")
        print(f"{'=' * 80}\n")

        print("DATASET CONTENTS:")
        print("─" * 80)
        print(f"📊 Main Dataset: {len(df)} countries × 45 years")
        print(f"📈 Summary Statistics: Comprehensive stats for each country")
        print(f"🌍 Regional Analysis: Time-series data by region")
        print(f"📅 Decade Analysis: Aggregated data by decade")
        print(f"🔥 Recent Trends: Focus on 2020-2024 period")
        print(f"📝 Metadata: Complete dataset documentation")
        print("─" * 80)

        print("\nPERFECT FOR:")
        print("  • Time series forecasting")
        print("  • Economic trend analysis")
        print("  • Machine learning models")
        print("  • Comparative studies")
        print("  • Data visualization projects")
        print("  • Statistical analysis")

        print(f"\n{'=' * 80}\n")

        return True

    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        print("Make sure you have 'openpyxl' installed:")
        print("  pip install openpyxl pandas numpy")
        return False


def main():
    """
    Main function
    """
    print("\n" + "=" * 80)
    print(" " * 20 + "INFLATION DATASET GENERATOR")
    print(" " * 15 + "For Data Science Projects")
    print("=" * 80)

    # Create dataset
    df = create_comprehensive_inflation_dataset()

    # Export to Excel
    success = export_to_excel(df)

    if success:
        print("✓ Your dataset is ready for analysis!")
        print("\nNEXT STEPS:")
        print("  1. Open the Excel file")
        print("  2. Explore different sheets")
        print("  3. Start your data science project!")
        print("\nHappy analyzing! 📊🚀\n")


if __name__ == "__main__":
    main()